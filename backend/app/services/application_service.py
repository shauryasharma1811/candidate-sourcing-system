"""
Applications business logic. Controllers call into this module only.
Covers both the admin review grid and the candidate-facing guided
application flow (Bio -> Education -> Experience -> Resume -> Review/Submit).
"""
import csv
import io
import math
import uuid
from datetime import date, datetime, timezone

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.common.enums import ApplicationStatus, NotificationEvent, ScanStatus
from app.core.config import settings
from app.models.application import Application
from app.models.notification import Notification, NotificationChannel, NotificationStatus
from app.models.user import User
from app.repositories.application_repository import ApplicationRepository
from app.repositories.job_repository import JobRepository
from app.repositories.notification_repository import NotificationRepository
from app.repositories.profile_repository import AdminRepository, CandidateRepository, EducationRepository, ExperienceRepository
from app.repositories.resume_repository import ResumeRepository
from app.services import notification_service, retention_service, storage_service, virus_scan_service
from app.schemas.application import (
    ApplicationDetail,
    ApplicationListItem,
    ApplicationStatusUpdateResponse,
    ResumeDownloadLink,
    TimelineEvent,
    experience_summary,
)
from app.schemas.candidate import (
    ApplicationProgressJob,
    ApplicationProgressResponse,
    ApplicationSubmitRequest,
    ApplicationSubmitResponse,
    BioDataResponse,
    EducationItem,
    ExperienceItem,
    MyApplicationItem,
    ProfilePhotoMetadata,
    ResumeMetadata,
)
from app.schemas.common import PaginatedMeta

MAX_PAGE_SIZE = 50

# Columns for the CSV/Excel export — kept as one ordered list so the header
# row and each data row can never drift apart.
EXPORT_COLUMNS = [
    "Application ID",
    "Candidate",
    "Email",
    "Job Title",
    "Requisition Code",
    "Applied On",
    "Experience",
    "Location",
    "Status",
]


class ApplicationService:
    def __init__(self, db: Session):
        self.db = db
        self.applications = ApplicationRepository(db)
        self.jobs = JobRepository(db)
        self.candidates = CandidateRepository(db)
        self.education = EducationRepository(db)
        self.experience = ExperienceRepository(db)
        self.resumes = ResumeRepository(db)
        self.admins = AdminRepository(db)
        self.notifications = NotificationRepository(db)

    def list_admin(
        self,
        page: int = 1,
        page_size: int = 20,
        status_filter: ApplicationStatus | None = None,
        job_id: str | None = None,
        search: str | None = None,
    ) -> tuple[list[ApplicationListItem], PaginatedMeta]:
        page = max(page, 1)
        page_size = min(max(page_size, 1), MAX_PAGE_SIZE)
        skip = (page - 1) * page_size
        job_uuid = self._parse_job_id(job_id)
        search = search.strip() if search and search.strip() else None

        items, total = self.applications.list_admin(
            skip=skip, limit=page_size, status=status_filter, job_id=job_uuid, search=search
        )

        data = [ApplicationListItem.from_model(app) for app in items]
        meta = PaginatedMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=max(1, math.ceil(total / page_size)),
        )
        return data, meta

    def _parse_job_id(self, job_id: str | None) -> uuid.UUID | None:
        if not job_id:
            return None
        try:
            return uuid.UUID(job_id)
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid job_id")

    def _get_admin_application_or_404(self, application_id: str) -> Application:
        try:
            app_uuid = uuid.UUID(application_id)
        except ValueError:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Application not found")

        application = self.applications.get_with_relations(app_uuid)
        if not application:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Application not found")
        return application

    def _build_timeline(self, application: Application) -> list[TimelineEvent]:
        """Submission, every notification dispatched off this application,
        and the current status/review, ordered oldest first — the
        candidate's full application review history in one list."""
        events: list[TimelineEvent] = [
            TimelineEvent(event="submitted", label="Application submitted", at=application.applied_at)
        ]

        # "Application Submitted" fans out to every admin (one Notification
        # row each) — collapse those into a single timeline entry keyed by
        # event type, keeping the earliest timestamp.
        seen_events: dict[str, TimelineEvent] = {}
        for notification in application.notifications or []:
            timestamp = notification.sent_at or notification.created_at
            if not timestamp:
                continue
            key = notification.event.value
            candidate_event = TimelineEvent(
                event=key,
                label=notification.subject or key.replace("_", " ").title(),
                at=timestamp,
                actor="admin" if notification.admin_recipient_id else "candidate",
            )
            if key not in seen_events or timestamp < seen_events[key].at:
                seen_events[key] = candidate_event
        events.extend(seen_events.values())

        if application.reviewed_at:
            reviewer = application.reviewed_by_admin
            events.append(
                TimelineEvent(
                    event="status_changed",
                    label=f"Status set to {application.status.value}",
                    at=application.reviewed_at,
                    actor=f"{reviewer.first_name} {reviewer.last_name}" if reviewer else None,
                )
            )

        events.sort(key=lambda e: e.at)
        return events

    def get_admin_detail(self, application_id: str) -> ApplicationDetail:
        """Backs 'View full application' — everything a recruiter needs to
        make a call, in one response."""
        application = self._get_admin_application_or_404(application_id)
        candidate = application.candidate
        job = application.job

        photo = None
        if candidate and candidate.photo_generated_filename:
            photo = ProfilePhotoMetadata(
                original_name=candidate.photo_original_name,
                mime_type=candidate.photo_mime_type,
                size_bytes=candidate.photo_size_bytes,
            )

        return ApplicationDetail(
            id=application.id,
            application_code=application.application_code,
            status=application.status,
            applied_at=application.applied_at,
            reviewed_at=application.reviewed_at,
            reviewed_by_admin_name=(
                f"{application.reviewed_by_admin.first_name} {application.reviewed_by_admin.last_name}"
                if application.reviewed_by_admin
                else None
            ),
            cover_note=application.cover_note,
            timeline=self._build_timeline(application),
            job=ApplicationProgressJob(
                id=job.id,
                title=job.title,
                requisition_code=job.requisition_code,
                department=job.department,
                location=job.location,
                status=job.status,
            ),
            bio=BioDataResponse(
                id=candidate.id,
                first_name=candidate.first_name,
                last_name=candidate.last_name,
                gender=candidate.gender,
                email=candidate.user.email if candidate.user else "",
                mobile=candidate.mobile,
                dob=candidate.dob,
                location=candidate.location,
                current_company=candidate.current_company,
                notice_period=candidate.notice_period,
                address=candidate.address,
                photo=photo,
            ),
            education=[EducationItem.model_validate(e) for e in candidate.education_entries],
            experience=[ExperienceItem.model_validate(e) for e in candidate.experience_entries],
            is_fresher=candidate.is_fresher,
            experience_summary=experience_summary(candidate),
            resume=ResumeMetadata.model_validate(application.resume) if application.resume else None,
        )

    def update_status(self, current_user: User, application_id: str, new_status: ApplicationStatus) -> ApplicationStatusUpdateResponse:
        """Inline status update from the grid. Records who reviewed it and
        when, and notifies the candidate on both channels: in-app (bell/
        list) and email — the BRD's 'future-ready' status-change event,
        now wired all the way through."""
        admin = self.admins.get_by_user_id(current_user.id)
        if not admin:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Admin profile not found")

        application = self._get_admin_application_or_404(application_id)

        application.status = new_status
        application.reviewed_by_admin_id = admin.id
        application.reviewed_at = datetime.now(timezone.utc)
        application = self.applications.update(application)

        if application.candidate:
            job_title = application.job.title if application.job else ""
            self.notifications.create(
                Notification(
                    event=NotificationEvent.STATUS_CHANGE,
                    channel=NotificationChannel.IN_APP,
                    status=NotificationStatus.SENT,
                    candidate_recipient_id=application.candidate_id,
                    application_id=application.id,
                    subject=f"Application status updated: {new_status.value}",
                    payload={"job_title": job_title, "status": new_status.value},
                )
            )

            # Status Change -> Candidate (email). Best-effort and
            # synchronous, same as the submission confirmation — a failed
            # send is recorded (channel=EMAIL, status=Failed) rather than
            # silently dropped, and never blocks the status update itself.
            candidate = application.candidate
            candidate_email = candidate.user.email if candidate.user else None
            if candidate_email:
                delivered = notification_service.send_status_change_email(
                    to_email=candidate_email,
                    candidate_name=f"{candidate.first_name} {candidate.last_name}",
                    job_title=job_title,
                    application_code=application.application_code,
                    new_status=new_status.value,
                )
                notification_service.record_notification(
                    self.db,
                    event=NotificationEvent.STATUS_CHANGE,
                    channel=NotificationChannel.EMAIL,
                    subject=f"Update on your application — {job_title}",
                    candidate_recipient_id=application.candidate_id,
                    application_id=application.id,
                    payload={"job_title": job_title, "status": new_status.value},
                    delivered=delivered,
                )

        return ApplicationStatusUpdateResponse(id=application.id, status=application.status, reviewed_at=application.reviewed_at)

    def get_resume_download_url(self, application_id: str) -> ResumeDownloadLink:
        application = self._get_admin_application_or_404(application_id)
        if not application.resume:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "No resume on file for this application")

        resume = application.resume
        # Secure downloads: never hand out a link for a file that hasn't
        # cleared the virus scan — infected/failed never got this far
        # (upload_resume rejects them outright), but a PENDING verdict
        # (the state a real async scanner would leave things in) must
        # also block download until it resolves.
        if resume.scan_status != ScanStatus.CLEAN:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "This resume hasn't cleared its security scan yet and can't be downloaded.",
            )

        url = storage_service.get_presigned_download_url(
            resume.generated_filename,
            resume.original_name,
            mime_type=resume.mime_type,
        )
        return ResumeDownloadLink(url=url, expires_in_seconds=settings.RESUME_DOWNLOAD_URL_TTL_SECONDS)

    def _export_rows(
        self,
        status_filter: ApplicationStatus | None = None,
        job_id: str | None = None,
        search: str | None = None,
    ) -> list[list[str]]:
        """Shared row-building for CSV and Excel export — same filters as the
        grid (status/job/search), so what a recruiter sees is exactly what
        they export, regardless of which format they pick."""
        job_uuid = self._parse_job_id(job_id)
        search = search.strip() if search and search.strip() else None
        applications = self.applications.list_for_export(status=status_filter, job_id=job_uuid, search=search)

        rows: list[list[str]] = []
        for application in applications:
            candidate = application.candidate
            job = application.job
            rows.append(
                [
                    application.application_code,
                    f"{candidate.first_name} {candidate.last_name}" if candidate else "",
                    candidate.user.email if candidate and candidate.user else "",
                    job.title if job else "",
                    job.requisition_code if job else "",
                    application.applied_at.strftime("%Y-%m-%d %H:%M") if application.applied_at else "",
                    experience_summary(candidate),
                    candidate.location if candidate else "",
                    application.status.value,
                ]
            )
        return rows

    def export_csv(
        self,
        status_filter: ApplicationStatus | None = None,
        job_id: str | None = None,
        search: str | None = None,
    ) -> str:
        """Export CSV — same filters as the grid (status/job/search), so
        what a recruiter sees is exactly what they export."""
        rows = self._export_rows(status_filter, job_id, search)
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(EXPORT_COLUMNS)
        writer.writerows(rows)
        return buffer.getvalue()

    def export_xlsx(
        self,
        status_filter: ApplicationStatus | None = None,
        job_id: str | None = None,
        search: str | None = None,
    ) -> bytes:
        """Export Excel (.xlsx) — M-1. BRD says "CSV/Excel"; this fills the
        gap flagged in the audit by reusing the exact same filtered rows as
        export_csv() so the two formats never drift from each other."""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        rows = self._export_rows(status_filter, job_id, search)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Applications"
        sheet.append(list(EXPORT_COLUMNS))
        for row in rows:
            sheet.append(row)

        # Freeze the header row and give columns a sane starting width so
        # the export is usable without a manual resize pass.
        sheet.freeze_panes = "A2"
        for idx, header in enumerate(EXPORT_COLUMNS, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 2)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Candidate-facing — the guided application flow
    # ------------------------------------------------------------------
    def _get_candidate_or_404(self, user: User):
        candidate = self.candidates.get_by_user_id(user.id)
        if not candidate:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Candidate profile not found")
        return candidate

    def _get_published_job_or_404(self, job_id: str):
        try:
            job_uuid = uuid.UUID(job_id)
        except ValueError:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Job not found")
        job = self.jobs.get_published_by_id(job_uuid)
        if not job:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Job not found")
        return job

    def get_progress(self, current_user: User, job_id: str) -> ApplicationProgressResponse:
        """Backs the Review & Submit step (and can be polled by earlier
        steps too) — one call returns the job plus everything the
        candidate has saved so far."""
        candidate = self._get_candidate_or_404(current_user)
        job = self._get_published_job_or_404(job_id)

        existing = self.applications.get_by_candidate_and_job(candidate.id, job.id)
        resume = self.resumes.get_latest_for_candidate(candidate.id)

        photo = None
        if candidate.photo_generated_filename:
            photo = ProfilePhotoMetadata(
                original_name=candidate.photo_original_name,
                mime_type=candidate.photo_mime_type,
                size_bytes=candidate.photo_size_bytes,
            )

        return ApplicationProgressResponse(
            job=ApplicationProgressJob(
                id=job.id,
                title=job.title,
                requisition_code=job.requisition_code,
                department=job.department,
                location=job.location,
                status=job.status,
            ),
            bio=BioDataResponse(
                id=candidate.id,
                first_name=candidate.first_name,
                last_name=candidate.last_name,
                gender=candidate.gender,
                email=current_user.email,
                mobile=candidate.mobile,
                dob=candidate.dob,
                location=candidate.location,
                current_company=candidate.current_company,
                notice_period=candidate.notice_period,
                address=candidate.address,
                photo=photo,
            ),
            education=[EducationItem.model_validate(e) for e in self.education.list_for_candidate(candidate.id)],
            experience=[ExperienceItem.model_validate(e) for e in self.experience.list_for_candidate(candidate.id)],
            resume=ResumeMetadata.model_validate(resume) if resume else None,
            is_fresher=candidate.is_fresher,
            already_applied=existing is not None,
            application_status=existing.status if existing else None,
        )

    def upload_resume(self, current_user: User, job_id: str, file: UploadFile) -> ResumeMetadata:
        candidate = self._get_candidate_or_404(current_user)
        self._get_published_job_or_404(job_id)  # 404s if the job isn't applicable

        original_name = file.filename or "resume"
        ext = "." + original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
        if ext not in settings.allowed_upload_extensions_list:
            allowed = ", ".join(settings.allowed_upload_extensions_list)
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Resume must be one of: {allowed}")

        # UploadFile doesn't expose size up front on every ASGI server —
        # read once, validate, and reuse the bytes for the scan + S3 upload.
        contents = file.file.read()
        size_bytes = len(contents)
        max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
        if size_bytes > max_bytes:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Resume must be {settings.MAX_UPLOAD_SIZE_MB}MB or smaller")
        if size_bytes == 0:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Resume file is empty")

        # Virus scan happens before anything touches S3 or the database —
        # an infected upload never gets stored anywhere, placeholder or not.
        scan_result = virus_scan_service.scan(contents, original_name)
        if scan_result.status == ScanStatus.INFECTED:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "This file failed our security scan and can't be uploaded.")
        if scan_result.status == ScanStatus.FAILED:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Couldn't scan this file. Please try again.")

        file.file.seek(0)
        object_key = storage_service.build_object_key(candidate.id, original_name)
        storage_service.upload_resume(file, object_key)

        from app.models.resume import Resume

        now = datetime.now(timezone.utc)
        resume = Resume(
            candidate_id=candidate.id,
            generated_filename=object_key,
            original_name=original_name,
            mime_type=file.content_type or "application/octet-stream",
            size_bytes=size_bytes,
            storage_bucket=settings.S3_BUCKET_NAME,
            scan_status=scan_result.status,
            scanned_at=now,
            scan_provider=scan_result.provider,
            retention_expires_at=retention_service.compute_retention_expiry(now),
        )
        resume = self.resumes.create(resume)
        return ResumeMetadata.model_validate(resume)

    def delete_resume(self, current_user: User, job_id: str) -> None:
        candidate = self._get_candidate_or_404(current_user)
        self._get_published_job_or_404(job_id)  # 404s if the job isn't applicable

        resume = self.resumes.get_latest_for_candidate(candidate.id)
        if not resume:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "No resume on file")

        object_key = resume.generated_filename
        self.resumes.delete(resume)
        storage_service.delete_resume(object_key)

    def _generate_application_code(self) -> str:
        """APP-<year>-<sequential>, e.g. APP-2026-00001. Mirrors
        JobService._generate_requisition_code: the sequence is derived from
        the current row count as a best guess, and the DB's unique
        constraint on application_code is the real guarantee against
        collisions — on the rare race we just retry with the next number."""
        year = date.today().year
        seq = self.applications.count_all() + 1
        while True:
            code = f"APP-{year}-{seq:05d}"
            if not self.applications.code_exists(code):
                return code
            seq += 1

    def _validate_submission(self, candidate) -> None:
        """Everything the BRD requires on file before an application can be
        submitted. Raises 400 on the first thing missing."""
        if not self.education.list_for_candidate(candidate.id):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Add at least one education entry before submitting")
        if not candidate.is_fresher and not self.experience.list_for_candidate(candidate.id):
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                'Add at least one work experience entry before submitting, or mark yourself as a Fresher',
            )
        if not candidate.mobile:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Complete your bio data before submitting")

    def submit_application(self, current_user: User, job_id: str, payload: ApplicationSubmitRequest) -> ApplicationSubmitResponse:
        candidate = self._get_candidate_or_404(current_user)
        job = self._get_published_job_or_404(job_id)

        # --- Validate everything ---
        if self.applications.get_by_candidate_and_job(candidate.id, job.id):
            raise HTTPException(status.HTTP_409_CONFLICT, "You've already applied to this job")

        self._validate_submission(candidate)

        resume = self.resumes.get_latest_for_candidate(candidate.id)
        if not resume:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload a resume before submitting")

        if not payload.consent:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Consent is required to submit an application")

        # --- Save application, linked to candidate + requisition, with a
        # generated Application ID. Retried once on a unique-constraint
        # race (duplicate application_code or a concurrent duplicate
        # submission for the same candidate/job) before giving up. ---
        for attempt in range(2):
            application = Application(
                application_code=self._generate_application_code(),
                candidate_id=candidate.id,
                job_id=job.id,
                resume_id=resume.id,
                status=ApplicationStatus.NEW,
                consent=payload.consent,
                cover_note=payload.cover_note,
            )
            try:
                application = self.applications.create(application)
                break
            except IntegrityError:
                self.db.rollback()
                if attempt == 1 or self.applications.get_by_candidate_and_job(candidate.id, job.id):
                    raise HTTPException(status.HTTP_409_CONFLICT, "You've already applied to this job")

        self._dispatch_submission_notifications(candidate, job, application)

        return ApplicationSubmitResponse(
            id=application.id,
            application_code=application.application_code,
            job_id=application.job_id,
            status=application.status,
            applied_at=application.applied_at,
        )

    def _dispatch_submission_notifications(self, candidate, job, application) -> None:
        """Application Submitted -> every Admin (in-app); Submission
        Confirmation -> the Candidate (in-app + email). Each channel is
        recorded as its own notifications row — the email row's status
        reflects the real SMTP result, not just the fact that the
        notification event happened."""
        for admin in self.admins.list_all():
            self.notifications.create(
                Notification(
                    event=NotificationEvent.APPLICATION_SUBMITTED,
                    channel=NotificationChannel.IN_APP,
                    status=NotificationStatus.SENT,
                    admin_recipient_id=admin.id,
                    application_id=application.id,
                    subject=f"New application: {job.title}",
                    payload={"job_title": job.title, "candidate_name": f"{candidate.first_name} {candidate.last_name}"},
                )
            )

        self.notifications.create(
            Notification(
                event=NotificationEvent.SUBMISSION_CONFIRMATION,
                channel=NotificationChannel.IN_APP,
                status=NotificationStatus.SENT,
                candidate_recipient_id=candidate.id,
                application_id=application.id,
                subject=f"Application received: {job.title}",
                payload={"job_title": job.title},
            )
        )

        candidate_email = candidate.user.email if candidate.user else None
        if candidate_email:
            delivered = notification_service.send_application_confirmation_email(
                to_email=candidate_email,
                candidate_name=f"{candidate.first_name} {candidate.last_name}",
                job_title=job.title,
                application_code=application.application_code,
            )
            notification_service.record_notification(
                self.db,
                event=NotificationEvent.SUBMISSION_CONFIRMATION,
                channel=NotificationChannel.EMAIL,
                subject=f"Application received — {job.title}",
                candidate_recipient_id=candidate.id,
                application_id=application.id,
                payload={"job_title": job.title},
                delivered=delivered,
            )

    def list_mine(self, current_user: User) -> list[MyApplicationItem]:
        candidate = self._get_candidate_or_404(current_user)
        items = self.applications.list_for_candidate(candidate.id)
        return [
            MyApplicationItem(
                id=app.id,
                application_code=app.application_code,
                job_id=app.job_id,
                job_title=app.job.title if app.job else "",
                status=app.status,
                applied_at=app.applied_at,
            )
            for app in items
        ]
